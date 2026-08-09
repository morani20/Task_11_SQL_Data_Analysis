import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Connect to the SQLite database
conn = sqlite3.connect('store.db')

print("--- 1. Executing SQL Analytical Queries ---")

# Q1: Top 5 customers by total spending (JOIN + GROUP BY + ORDER BY)
q1_sql = '''
SELECT c.customer_id, c.customer_name, c.city, SUM(s.quantity * p.price) AS total_spent
FROM customers c
JOIN sales s ON c.customer_id = s.customer_id
JOIN products p ON s.product_id = p.product_id
GROUP BY c.customer_id
ORDER BY total_spent DESC
LIMIT 5;
'''
df_top_customers = pd.read_sql_query(q1_sql, conn)

# Q2: Customers with zero purchases (LEFT JOIN + WHERE ... IS NULL)
q2_sql = '''
SELECT c.customer_id, c.customer_name, c.city, c.signup_date
FROM customers c
LEFT JOIN sales s ON c.customer_id = s.customer_id
WHERE s.sale_id IS NULL;
'''
df_no_purchase = pd.read_sql_query(q2_sql, conn)

# Q3: Customer average spending vs overall average spending (Subquery)
q3_sql = '''
SELECT c.customer_name, 
       ROUND(AVG(s.quantity * p.price), 2) AS customer_avg_spent,
       ROUND((SELECT AVG(s2.quantity * p2.price) 
              FROM sales s2 
              JOIN products p2 ON s2.product_id = p2.product_id), 2) AS overall_avg_spent
FROM customers c
JOIN sales s ON c.customer_id = s.customer_id
JOIN products p ON s.product_id = p.product_id
GROUP BY c.customer_id;
'''
df_avg_comparison = pd.read_sql_query(q3_sql, conn)

# Q4: Customer ranking by city based on spend (Window Function)
q4_sql = '''
SELECT c.city, c.customer_name, SUM(s.quantity * p.price) AS total_spent,
       DENSE_RANK() OVER (PARTITION BY c.city ORDER BY SUM(s.quantity * p.price) DESC) as city_rank
FROM customers c
JOIN sales s ON c.customer_id = s.customer_id
JOIN products p ON s.product_id = p.product_id
GROUP BY c.customer_id, c.city;
'''
df_city_ranking = pd.read_sql_query(q4_sql, conn)

# Q5: Monthly revenue trend (strftime + GROUP BY)
q5_sql = '''
SELECT strftime('%Y-%m', s.sale_date) AS month, SUM(s.quantity * p.price) AS monthly_revenue
FROM sales s
JOIN products p ON s.product_id = p.product_id
GROUP BY month
ORDER BY month;
'''
df_monthly_trend = pd.read_sql_query(q5_sql, conn)

# Q6: Product category performance by revenue and volume
q6_sql = '''
SELECT p.category, SUM(s.quantity * p.price) AS category_revenue, SUM(s.quantity) AS total_units_sold
FROM sales s
JOIN products p ON s.product_id = p.product_id
GROUP BY p.category
ORDER BY category_revenue DESC;
'''
df_category_revenue = pd.read_sql_query(q6_sql, conn)

# Q7: Signup date vs total spending
q7_sql = '''
SELECT c.customer_name, c.signup_date, SUM(s.quantity * p.price) AS total_spent
FROM customers c
JOIN sales s ON c.customer_id = s.customer_id
JOIN products p ON s.product_id = p.product_id
GROUP BY c.customer_id
ORDER BY c.signup_date ASC;
'''
df_signup_vs_spent = pd.read_sql_query(q7_sql, conn)

# Q8: Percentage of repeat customers (>1 purchase)
q8_sql = '''
WITH CustomerOrders AS (
    SELECT customer_id, COUNT(sale_id) as order_count
    FROM sales
    GROUP BY customer_id
)
SELECT 
    ROUND(COUNT(CASE WHEN order_count > 1 THEN 1 END) * 100.0 / (SELECT COUNT(*) FROM customers), 2) AS repeat_customer_percentage
FROM CustomerOrders;
'''
df_repeat_ratio = pd.read_sql_query(q8_sql, conn)

conn.close()

print("--- 2. Generating Visualizations ---")
sns.set_theme(style="whitegrid")

# Chart 1: Top 5 Customers (Matplotlib / Seaborn)
plt.figure(figsize=(8, 5))
sns.barplot(data=df_top_customers, x='customer_name', y='total_spent', hue='customer_name', palette='Blues_r', legend=False)
plt.title('Top 5 Spending Customers', fontsize=14, fontweight='bold', pad=15)
plt.xlabel('Customer Name', fontweight='bold')
plt.ylabel('Total Spent ($)', fontweight='bold')
plt.xticks(rotation=15)
plt.tight_layout()
plt.savefig('chart_top_customers.png', dpi=300)
plt.close()

# Chart 2: Monthly Revenue Trend
plt.figure(figsize=(8, 5))
plt.plot(df_monthly_trend['month'], df_monthly_trend['monthly_revenue'], marker='o', color='#2ca02c', linewidth=2.5)
plt.title('Monthly Revenue Trend', fontsize=14, fontweight='bold', pad=15)
plt.xlabel('Month', fontweight='bold')
plt.ylabel('Revenue ($)', fontweight='bold')
plt.grid(True, linestyle='--', alpha=0.6)
plt.tight_layout()
plt.savefig('chart_monthly_trend.png', dpi=300)
plt.close()

# Chart 3: Revenue by Category
plt.figure(figsize=(8, 5))
sns.barplot(data=df_category_revenue, x='category', y='category_revenue', hue='category', palette='viridis', legend=False)
plt.title('Revenue by Product Category', fontsize=14, fontweight='bold', pad=15)
plt.xlabel('Category', fontweight='bold')
plt.ylabel('Revenue ($)', fontweight='bold')
plt.tight_layout()
plt.savefig('chart_category_revenue.png', dpi=300)
plt.close()

# Chart 4: Interactive Plotly Chart (Signup Date vs Spending)
fig_plotly = px.scatter(
    df_signup_vs_spent, 
    x='signup_date', 
    y='total_spent', 
    size='total_spent', 
    color='customer_name',
    title='Interactive: Customer Signup Date vs Total Spending',
    labels={'signup_date': 'Signup Date', 'total_spent': 'Total Spent ($)'}
)
fig_plotly.write_html('chart_signup_vs_spending_interactive.html')

print("--- 3. Exporting Professional Excel Report ---")

# Initialize OpenPyXL workbook
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

# Styling configurations
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy Blue
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
kpi_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue
alert_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Light Orange
zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Alternating rows
thin_border = Border(left=Side(style='thin', color='D9D9D9'),
                     right=Side(style='thin', color='D9D9D9'),
                     top=Side(style='thin', color='D9D9D9'),
                     bottom=Side(style='thin', color='D9D9D9'))

title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")

def create_styled_sheet(wb, title, df):
    ws = wb.create_sheet(title=title)
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.merge_cells("A1:E1")
    ws["A1"] = title.upper()
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35
    
    start_row = 3
    
    # Write Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = col_name.replace('_', ' ').title()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 26
    
    # Write Data
    for row_idx, row_data in enumerate(df.values, start=start_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border
            
            # Alternating row fill
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
                
            # Number formatting
            col_name = df.columns[col_idx-1]
            if isinstance(val, (int, float)):
                if any(k in col_name for k in ['spent', 'revenue', 'price', 'avg']):
                    cell.number_format = '$#,##0.00'
                elif isinstance(val, float):
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

# 1. Summary Dashboard Sheet
ws_sum = wb.create_sheet(title="Summary", index=0)
ws_sum.views.sheetView[0].showGridLines = True

ws_sum["A1"] = "EXECUTIVE STORE ANALYSIS DASHBOARD"
ws_sum["A1"].font = title_font

# KPI 1: Total Customers
ws_sum.merge_cells("A3:B3")
ws_sum["A3"] = "TOTAL CUSTOMERS"
ws_sum["A3"].font = Font(name="Calibri", size=10, bold=True, color="595959")
ws_sum["A3"].fill = kpi_fill
ws_sum["A3"].alignment = Alignment(horizontal="center")

ws_sum.merge_cells("A4:B4")
ws_sum["A4"] = 15
ws_sum["A4"].font = Font(name="Calibri", size=18, bold=True, color="1F4E78")
ws_sum["A4"].alignment = Alignment(horizontal="center")

# KPI 2: Repeat Customer Rate
ws_sum.merge_cells("D3:E3")
ws_sum["D3"] = "REPEAT CUSTOMER RATE"
ws_sum["D3"].font = Font(name="Calibri", size=10, bold=True, color="595959")
ws_sum["D3"].fill = kpi_fill
ws_sum["D3"].alignment = Alignment(horizontal="center")

ws_sum.merge_cells("D4:E4")
ws_sum["D4"] = f"{df_repeat_ratio.iloc[0,0]}%"
ws_sum["D4"].font = Font(name="Calibri", size=18, bold=True, color="1F4E78")
ws_sum["D4"].alignment = Alignment(horizontal="center")

# KPI 3: Inactive Customers
ws_sum.merge_cells("G3:H3")
ws_sum["G3"] = "INACTIVE CUSTOMERS"
ws_sum["G3"].font = Font(name="Calibri", size=10, bold=True, color="595959")
ws_sum["G3"].fill = alert_fill
ws_sum["G3"].alignment = Alignment(horizontal="center")

ws_sum.merge_cells("G4:H4")
ws_sum["G4"] = len(df_no_purchase)
ws_sum["G4"].font = Font(name="Calibri", size=18, bold=True, color="C00000")
ws_sum["G4"].alignment = Alignment(horizontal="center")

# Overview Table on Summary Sheet
ws_sum["A7"] = "Key Metrics Overview"
ws_sum["A7"].font = Font(name="Calibri", size=13, bold=True, color="1F4E78")

summary_table = [
    ["Metric Description", "Value"],
    ["Top Spender", df_top_customers.iloc[0]['customer_name']],
    ["Top Category Revenue", f"${df_category_revenue.iloc[0]['category_revenue']:,.2f}"],
    ["Overall Average Purchase", f"${df_avg_comparison.iloc[0]['overall_avg_spent']:,.2f}"]
]

for r_idx, row in enumerate(summary_table, start=8):
    for c_idx, val in enumerate(row, start=1):
        cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        if r_idx == 8:
            cell.fill = header_fill
            cell.font = header_font
        else:
            if r_idx % 2 == 0:
                cell.fill = zebra_fill

for col in ws_sum.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_sum.column_dimensions[col_letter].width = max(max_len + 3, 15)

# 2. Add sub-sheets
create_styled_sheet(wb, "Top 5 Customers", df_top_customers)
create_styled_sheet(wb, "Inactive Customers", df_no_purchase)
create_styled_sheet(wb, "Avg Spend Comparison", df_avg_comparison)
create_styled_sheet(wb, "City Rankings", df_city_ranking)
create_styled_sheet(wb, "Monthly Trend", df_monthly_trend)
create_styled_sheet(wb, "Category Performance", df_category_revenue)

# Save workbook
wb.save('customer_analysis_report.xlsx')
print("✓ customer_analysis_report.xlsx successfully exported with corporate styling!")